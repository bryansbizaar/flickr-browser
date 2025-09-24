#!/usr/bin/env python3
"""
Excel Exporter for Flickr Local Browser
Creates Excel spreadsheet with thumbnails and metadata from local database
"""

import sqlite3
import os
import sys
from pathlib import Path
from PIL import Image
import io
import openpyxl
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.worksheet.dimensions import ColumnDimension
import pandas as pd
from datetime import datetime

class FlickrExcelExporter:
    def __init__(self, data_dir="data"):
        """Initialize the exporter with database and thumbnail paths"""
        if not os.path.isabs(data_dir):
            data_dir = os.path.abspath(data_dir)
        
        self.data_dir = Path(data_dir)
        self.db_path = self.data_dir / "flickr_metadata.db"
        self.thumbnails_dir = self.data_dir / "thumbnails"
        
        print(f"Database path: {self.db_path}")
        print(f"Thumbnails path: {self.thumbnails_dir}")
        
        if not self.db_path.exists():
            raise FileNotFoundError(f"Database not found at {self.db_path}")
        
        if not self.thumbnails_dir.exists():
            raise FileNotFoundError(f"Thumbnails directory not found at {self.thumbnails_dir}")

    def get_db_connection(self):
        """Get database connection with Row factory for named columns"""
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        return conn

    def get_all_photos_with_metadata(self):
        """Extract all photos with complete metadata including album associations"""
        conn = self.get_db_connection()
        cursor = conn.cursor()
        
        print("Extracting photo metadata from database...")
        
        # Main query to get all photo data with album associations
        query = """
        SELECT 
            p.id,
            p.title,
            p.description,
            p.tags,
            p.date_taken,
            p.date_posted,
            p.views,
            p.url_thumbnail,
            p.url_original,
            p.owner_username,
            p.owner_realname,
            p.latitude,
            p.longitude,
            p.accuracy,
            p.media_type,
            p.original_secret,
            p.original_format,
            p.can_comment,
            p.can_print,
            p.can_share,
            p.can_blog,
            p.can_download,
            p.rotation,
            p.public,
            p.friend,
            p.family,
            p.safety_level,
            p.license,
            p.path_alias,
            p.album_id as primary_album_id,
            GROUP_CONCAT(a.title, '; ') as all_albums,
            COUNT(c.id) as comment_count
        FROM photos p
        LEFT JOIN photo_albums pa ON p.id = pa.photo_id
        LEFT JOIN albums a ON pa.album_id = a.id
        LEFT JOIN comments c ON p.id = c.photo_id
        GROUP BY p.id
        ORDER BY p.date_taken DESC
        """
        
        cursor.execute(query)
        photos = [dict(row) for row in cursor.fetchall()]
        
        conn.close()
        print(f"Found {len(photos)} photos in database")
        return photos

    def prepare_thumbnail_image(self, photo_id, target_size=(100, 100)):
        """Prepare thumbnail image for Excel embedding"""
        thumbnail_path = self.thumbnails_dir / f"{photo_id}.jpg"
        
        if not thumbnail_path.exists():
            print(f"Warning: Thumbnail not found for photo {photo_id}")
            return None
        
        try:
            # Open and resize image
            with Image.open(thumbnail_path) as img:
                # Convert to RGB if necessary (for JPEG compatibility)
                if img.mode != 'RGB':
                    img = img.convert('RGB')
                
                # Resize while maintaining aspect ratio
                img.thumbnail(target_size, Image.Resampling.LANCZOS)
                
                # Save to BytesIO buffer
                img_buffer = io.BytesIO()
                img.save(img_buffer, format='JPEG', quality=85)
                img_buffer.seek(0)
                
                return img_buffer
                
        except Exception as e:
            print(f"Error processing thumbnail for photo {photo_id}: {e}")
            return None

    def create_excel_workbook(self, photos, include_thumbnails=True, max_rows=None):
        """Create Excel workbook with photo data"""
        print(f"Creating Excel workbook with {len(photos)} photos...")
        
        # Limit rows if specified (for testing or large datasets)
        if max_rows and len(photos) > max_rows:
            photos = photos[:max_rows]
            print(f"Limited to first {max_rows} photos for manageable file size")
        
        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Flickr Photos"
        
        # Define columns
        columns = [
            ('Thumbnail', 15),
            ('Title', 30),
            ('Description', 40), 
            ('Tags', 30),
            ('Date Taken', 15),
            ('Date Posted', 15),
            ('Views', 8),
            ('Albums', 25),
            ('Comments', 8),
            ('Owner', 15),
            ('Real Name', 15),
            ('Public', 8),
            ('Family', 8),
            ('Friends', 8),
            ('Media Type', 10),
            ('Format', 8),
            ('Latitude', 12),
            ('Longitude', 12),
            ('Safety Level', 10),
            ('License', 8),
            ('Can Comment', 10),
            ('Can Download', 10),
            ('Photo ID', 12)
        ]
        
        # Set up headers
        for col_idx, (header, width) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Set column width
            ws.column_dimensions[chr(64 + col_idx)].width = width
        
        # Set header row height
        ws.row_dimensions[1].height = 25
        
        print("Adding photo data to Excel...")
        
        # Add photo data
        for row_idx, photo in enumerate(photos, 2):
            # Set row height for thumbnails
            if include_thumbnails:
                ws.row_dimensions[row_idx].height = 80
            
            # Helper function to safely get values
            def safe_get(key, default=''):
                value = photo.get(key)
                if value is None:
                    return default
                return str(value)
            
            # Format dates
            date_taken = photo.get('date_taken', '')
            date_posted = photo.get('date_posted', '')
            
            try:
                if date_taken:
                    date_taken = datetime.fromisoformat(date_taken.replace('Z', '+00:00')).strftime('%Y-%m-%d %H:%M')
                if date_posted:
                    date_posted = datetime.fromisoformat(date_posted.replace('Z', '+00:00')).strftime('%Y-%m-%d %H:%M')
            except:
                pass  # Keep original format if parsing fails
            
            # Add data to row
            row_data = [
                '',  # Thumbnail placeholder
                safe_get('title'),
                safe_get('description'),
                safe_get('tags'),
                date_taken,
                date_posted,
                safe_get('views', '0'),
                safe_get('all_albums'),
                safe_get('comment_count', '0'),
                safe_get('owner_username'),
                safe_get('owner_realname'),
                'Yes' if photo.get('public') == 1 else 'No',
                'Yes' if photo.get('family') == 1 else 'No',
                'Yes' if photo.get('friend') == 1 else 'No',
                safe_get('media_type'),
                safe_get('original_format'),
                safe_get('latitude'),
                safe_get('longitude'),
                safe_get('safety_level'),
                safe_get('license'),
                'Yes' if photo.get('can_comment') == 1 else 'No',
                'Yes' if photo.get('can_download') == 1 else 'No',
                safe_get('id')
            ]
            
            # Add text data to cells
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(vertical='center', wrap_text=True)
            
            # Add thumbnail image if requested
            if include_thumbnails:
                img_buffer = self.prepare_thumbnail_image(photo['id'])
                if img_buffer:
                    try:
                        img = ExcelImage(img_buffer)
                        # Position image in the thumbnail column
                        img.anchor = f"A{row_idx}"
                        ws.add_image(img)
                    except Exception as e:
                        print(f"Warning: Could not add thumbnail for photo {photo['id']}: {e}")
            
            # Progress indicator
            if row_idx % 100 == 0:
                print(f"Processed {row_idx - 1} photos...")
        
        # Apply autofilter to all columns
        ws.auto_filter.ref = ws.dimensions
        
        # Freeze header row and thumbnail column
        ws.freeze_panes = 'B2'
        
        return wb

    def export_to_excel(self, output_file=None, include_thumbnails=True, max_rows=None):
        """Main export function"""
        if output_file is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = f"flickr_photos_export_{timestamp}.xlsx"
        
        print(f"\n=== Flickr Excel Export ===")
        print(f"Output file: {output_file}")
        print(f"Include thumbnails: {include_thumbnails}")
        if max_rows:
            print(f"Max rows: {max_rows}")
        print()
        
        try:
            # Get photo data
            photos = self.get_all_photos_with_metadata()
            
            if not photos:
                print("No photos found in database!")
                return
            
            # Create Excel workbook
            wb = self.create_excel_workbook(photos, include_thumbnails, max_rows)
            
            # Save file
            print(f"Saving Excel file: {output_file}")
            wb.save(output_file)
            
            file_size = os.path.getsize(output_file) / (1024 * 1024)  # Convert to MB
            print(f"\n✅ Export complete!")
            print(f"📄 File: {output_file}")
            print(f"📊 Photos: {len(photos) if not max_rows else min(len(photos), max_rows)}")
            print(f"💾 Size: {file_size:.1f} MB")
            
        except Exception as e:
            print(f"\n❌ Export failed: {e}")
            raise

def main():
    """Command line interface"""
    import argparse
    
    parser = argparse.ArgumentParser(description='Export Flickr photos to Excel')
    parser.add_argument('--output', '-o', help='Output Excel file name')
    parser.add_argument('--no-thumbnails', action='store_true', 
                       help='Skip thumbnail images (faster, smaller file)')
    parser.add_argument('--max-rows', type=int, 
                       help='Limit number of photos (for testing)')
    parser.add_argument('--data-dir', default='../data',
                       help='Path to data directory (default: ../data)')
    
    args = parser.parse_args()
    
    try:
        exporter = FlickrExcelExporter(args.data_dir)
        exporter.export_to_excel(
            output_file=args.output,
            include_thumbnails=not args.no_thumbnails,
            max_rows=args.max_rows
        )
        
    except Exception as e:
        print(f"Error: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()
